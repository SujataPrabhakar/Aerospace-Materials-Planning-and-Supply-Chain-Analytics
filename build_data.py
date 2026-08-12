"""Build the compact JSON model that backs the Aerospace Materials Planning dashboard.

Every field is sourced from Aerospace_Supply_Chain_project.xlsx / the source extracts.
The only DERIVED values are the 3-month moving-average backtest MAE per Part-Site
(same method as the workbook's own Demand_Forcast sheet, extended to all combinations)
and the client-side re-aggregation of the supplier PO metrics, whose definitions were
validated against the workbook's Sheet14 grand totals to 6 decimal places.
"""
import json
import math
import pandas as pd
import numpy as np
import openpyxl

BOOK = "Aerospace_Supply_Chain_project.xlsx"

# ---------------------------------------------------------------- forecast sheet
wb = openpyxl.load_workbook(BOOK, read_only=True, data_only=True)
ws = wb["All_Part_Site_Forecast"]
rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]

HIST = [f"{y}-{m:02d}" for y in (2022, 2023, 2024) for m in range(1, 13)]
FCST = [f"2025-{m:02d}" for m in range(1, 13)]

# map header position -> month label (col 37 'Aprl-25' is a text typo for 2025-04)
col_hist, col_fc = {}, {}
for i, h in enumerate(hdr):
    if hasattr(h, "year"):
        lbl = f"{h.year}-{h.month:02d}"
        (col_hist if h.year < 2025 else col_fc)[lbl] = i
    elif isinstance(h, str) and h.strip() == "Aprl-25":
        col_fc["2025-04"] = i
hi = [col_hist[m] for m in HIST]
fi = [col_fc[m] for m in FCST]
L = {str(h).strip(): i for i, h in enumerate(hdr) if isinstance(h, str)}
i_ftot, i_atot, i_chg = L["2025 Forecast Total"], L["2024 Actual Total"], L["Forecast Change %"]

fc = {}
for r in rows[1:]:
    if not r or not r[0]:
        continue
    fc[r[0]] = dict(
        hist=[float(r[c] or 0) for c in hi],
        fcst=[round(float(r[c] or 0), 2) for c in fi],
        ftot=float(r[i_ftot] or 0),
        atot=float(r[i_atot] or 0),
        chg=float(r[i_chg]) if r[i_chg] is not None else None,
    )

# ---------------------------------------------------------------- inventory plan
ws = wb["Inventory_Planning"]
inv_rows = list(ws.iter_rows(values_only=True))
wb.close()
inv_hdr = list(inv_rows[0])
IX = {h: i for i, h in enumerate(inv_hdr)}
inv = {}
for r in inv_rows[1:]:
    if not r or not r[0]:
        continue
    raw = lambda k: r[IX[k]]

    def g(k):
        """Blank / 'N/A' cells (zero-forecast items) collapse to 0."""
        v = raw(k)
        return 0 if isinstance(v, str) and v.strip().upper() in ("N/A", "#DIV/0!", "") else v

    inv[r[0]] = dict(
        blocked=float(g("Sum of blocked_qty") or 0),
        backorder=float(g("Sum of backorder_qty") or 0),
        onhand=float(g("Sum of on_hand_qty") or 0),
        usable=float(g("Usable Inventory") or 0),
        avgfc=float(g("Avg Monthly Forecast") or 0),
        cover=float(g("Months of Cover") or 0),
        lt=float(g("Lead Time Days") or 0),
        ltd=float(g("Lead Time Demand") or 0),
        crit=str(raw("Criticality") or ""),
        tsl=float(g("Target Service Level") or 0),
        std=float(g("Monthly Demand Std Dev") or 0),
        z=float(g("Z-Score") or 0),
        ss=float(g("Safety Stock") or 0),
        rop=float(g("Reorde Point") or 0),
        duein=float(g("Due In Qty") or 0),
        pos=float(g("Inventory Position") or 0),
        status=str(raw("Replenishment Status") or ""),
        gap=float(g("Reorder Gap") or 0),
        tgt=float(g("Target Stock Level") or 0),
        rec=float(g("Recommended Order Qty") or 0),
    )

# ---------------------------------------------------------------- masters + POs
pm = pd.read_excel("Parts_Master.xlsx").set_index("part_id")
po = pd.read_excel("Purchase_Orders.xlsx").merge(
    pm[["lead_time_days"]], left_on="part_id", right_index=True, how="left"
)
po["delay"] = (po.receipt_date - po.promised_date).dt.days
po["ontime"] = (po.delay <= 0).astype(int)
po["actual_lt"] = (po.receipt_date - po.order_date).dt.days
po["ltvar"] = po.actual_lt - po.lead_time_days
po["short"] = (po.ordered_qty - po.received_qty).clip(lower=0)
po["late_only"] = po.delay.where(po.delay > 0)
po["key"] = po.part_id + "-" + po.site_id

agg = po.groupby("key").agg(
    po_n=("po_id", "size"),
    po_ontime=("ontime", "sum"),
    late_sum=("late_only", "sum"),
    late_n=("late_only", "count"),
    alt_sum=("actual_lt", "sum"),
    var_sum=("ltvar", "sum"),
    ordered=("ordered_qty", "sum"),
    received=("received_qty", "sum"),
    shortfall=("short", "sum"),
)

qi = pd.read_excel("Quality_Incidents.xlsx")
qi["key"] = qi.part_id + "-" + qi.site_id
qagg = qi.groupby("key").agg(inc=("incident_id", "size"), scrap=("scrap_qty", "sum"))


def backtest_mae(h):
    """3-month moving-average backtest, identical method to the Demand_Forcast sheet."""
    errs = [abs(h[t] - sum(h[t - 3:t]) / 3) for t in range(3, len(h))]
    return round(float(np.mean(errs)), 2)


# ---------------------------------------------------------------- assemble rows
COLS = ["part", "site", "family", "crit", "sup", "onhand", "blocked", "backorder",
        "usable", "duein", "pos", "ftot", "avgfc", "cover", "lt", "ltd", "tsl",
        "std", "ss", "rop", "status", "gap", "tgt", "rec", "atot", "chg", "cost",
        "risk", "po_n", "po_ontime", "late_sum", "late_n", "alt_sum", "var_sum",
        "ordered", "received", "shortfall", "mae", "inc", "scrap", "hist", "fcst"]

out = []
for key, iv in inv.items():
    part, site = key.split("-")
    f = fc.get(key)
    if f is None:
        continue
    m = pm.loc[part]
    a = agg.loc[key] if key in agg.index else None
    q = qagg.loc[key] if key in qagg.index else None
    out.append([
        part, site, str(m.part_family), iv["crit"], str(m.supplier_id_primary),
        iv["onhand"], iv["blocked"], iv["backorder"], iv["usable"], iv["duein"], iv["pos"],
        round(f["ftot"], 2), round(iv["avgfc"], 2), round(iv["cover"], 2), iv["lt"],
        round(iv["ltd"], 2), iv["tsl"], iv["std"], iv["ss"], iv["rop"], iv["status"],
        iv["gap"], iv["tgt"], iv["rec"], f["atot"],
        round(f["chg"], 4) if f["chg"] is not None else None,
        float(m.unit_cost), str(m.supplier_risk_class),
        int(a.po_n) if a is not None else 0,
        int(a.po_ontime) if a is not None else 0,
        float(a.late_sum) if a is not None and not math.isnan(a.late_sum) else 0.0,
        int(a.late_n) if a is not None else 0,
        float(a.alt_sum) if a is not None else 0.0,
        float(a.var_sum) if a is not None else 0.0,
        int(a.ordered) if a is not None else 0,
        int(a.received) if a is not None else 0,
        int(a.shortfall) if a is not None else 0,
        backtest_mae(f["hist"]),
        int(q.inc) if q is not None else 0,
        int(q.scrap) if q is not None else 0,
        [int(x) for x in f["hist"]],
        f["fcst"],
    ])

model = dict(
    cols=COLS,
    histMonths=HIST,
    fcstMonths=FCST,
    rows=out,
    validation=dict(
        otd=0.44151553967504886, fill=0.9715619313594004, shortfall=12362,
        avgLate=3.329249154997586, avgActualLT=42.9715162138475,
        ordered=434699, received=422337, mae_p1s1=1.85,
    ),
)
with open("dashboard_data.json", "w") as fh:
    json.dump(model, fh, separators=(",", ":"))

# ---------------------------------------------------------------- sanity checks
import os
print("rows:", len(out), "| json MB:", round(os.path.getsize("dashboard_data.json") / 1e6, 2))
d = {r[0] + "-" + r[1]: r for r in out}
print("MAE P00001-SITE01:", d["P00001-SITE01"][COLS.index("mae")], "(sheet says 1.85)")
tot = lambda c: sum(r[COLS.index(c)] for r in out)
print("OTD:", tot("po_ontime") / tot("po_n"), "| fill:", tot("received") / tot("ordered"))
print("shortfall:", tot("shortfall"), "| avg late:", tot("late_sum") / tot("late_n"))
print("avg actual LT:", tot("alt_sum") / tot("po_n"), "| avg var:", tot("var_sum") / tot("po_n"))
print("reorder rows:", sum(1 for r in out if r[COLS.index("status")] == "Reorder"))
print("statuses:", {r[COLS.index("status")] for r in out})
print("2025 fcst:", round(tot("ftot")), "| 2024 act:", tot("atot"),
      "| chg%:", round(100 * (tot("ftot") / tot("atot") - 1), 2))
print("usable:", tot("usable"), "| duein:", tot("duein"), "| recQty:", tot("rec"))
print("cover:", round(tot("usable") / tot("avgfc"), 2))
